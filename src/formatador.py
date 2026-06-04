from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


def formatar_excel(caminho_arquivo):

    wb = load_workbook(caminho_arquivo)

    for aba in wb.worksheets:

        # Formatação do cabeçalho
        for celula in aba[1]:

            celula.font = Font(
                bold=True,
                color="FFFFFF"
            )

            celula.fill = PatternFill(
                fill_type="solid",
                fgColor="1F4E78"
            )

            celula.alignment = Alignment(
                horizontal="center"
            )

        # Ajuste automático das colunas
        for coluna in aba.columns:

            tamanho_maximo = 0
            letra_coluna = coluna[0].column_letter

            for celula in coluna:

                try:
                    if celula.value is not None:
                        tamanho_maximo = max(
                            tamanho_maximo,
                            len(str(celula.value))
                        )
                except:
                    pass

            aba.column_dimensions[
                letra_coluna
            ].width = tamanho_maximo + 3

    wb.save(caminho_arquivo)