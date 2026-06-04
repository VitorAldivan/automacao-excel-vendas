from openpyxl import load_workbook
from openpyxl.styles import Font


def criar_dashboard(
    caminho_arquivo,
    kpis
):

    wb = load_workbook(
        caminho_arquivo
    )

    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]

    dashboard = wb.create_sheet(
        title="Dashboard",
        index=0
    )

    dashboard["A1"] = (
        "DASHBOARD DE VENDAS"
    )

    dashboard["A1"].font = Font(
        bold=True,
        size=16
    )

    linha = 3

    for _, row in kpis.iterrows():

        dashboard[f"A{linha}"] = (
            row["Indicador"]
        )

        dashboard[f"B{linha}"] = (
            row["Valor"]
        )

        linha += 1

    wb.save(
        caminho_arquivo
    )